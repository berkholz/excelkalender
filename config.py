

############################ IMPORTS #####################
from openpyxl import Workbook
import datetime
import calendar
from calendar import monthrange
from openpyxl.styles import PatternFill,Font, Color, Border, Side, Alignment

############################ VARIABLES #####################
excel_file_name = "kalender.xlsx"

users = ['Marcel','Antje','Johannes','Irma','Arthur']

cal = calendar.Calendar()
#actual_year = datetime.date.today().year
actual_year = 2023

def add_header(worksheet):
    worksheet.merge_cells('A1:AG1')
    worksheet.row_dimensions[1].height = 30
    cell = ws.cell(row=1,column=1,value=actual_year)
    cell.font = Font(name='Arial',size=16,bold=True)
    cell.alignment = Alignment(horizontal='center',vertical='center')
    cell.border = Border(bottom=Side(border_style="thin", color="000000"))
    #cell.fill = PatternFill(start_color="CFCFCF", end_color="0F0F0F", fill_type="solid")

def set_header_colomn_style (cell):
    cell.font = Font(name='Arial', bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.border = Border(bottom=Side(border_style="thin", color="000000"))
    cell.fill = PatternFill(start_color="CFCFCF", end_color="0F0F0F", fill_type="solid")

def set_column_width(workspace):
    # set cell width of month name
    column_range = ['A','AG']
    for column in column_range:
        workspace.column_dimensions[column].width = 25

    # set cell width of days of month
    column_range = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF']
    for column in column_range:
        workspace.column_dimensions[column].width = 5

############################ MAIN #####################
wb = Workbook()

# grab the active worksheet
ws = wb.active
ws.title = str(actual_year)

add_header(ws)
set_column_width(ws)

row_count = 3

### iterating over the last month of the year before
_, last_day_of_month = monthrange(actual_year - 1, 12)
actual_cell = ws.cell(row=row_count,column=1,value=calendar.month_name[12])
set_header_colomn_style(actual_cell)

column_position = 0
for column in range(1,32):
    if column_position < last_day_of_month:
        actual_cell = ws.cell(row=row_count, column=column+1, value=column)
        set_header_colomn_style(actual_cell)
    else:
        actual_cell = ws.cell(row=row_count, column=column+1, value='')
        set_header_colomn_style(actual_cell)
    column_position += 1
actual_cell = ws.cell(row=row_count,column=33,value=calendar.month_name[12])
set_header_colomn_style(actual_cell)
row_count = row_count + 1
for user in users:
    ws.cell(row=row_count,column=1,value=user)
    ws.cell(row=row_count,column=33,value=user)
    row_count = row_count + 1
# let two rows empty for spacing
row_count = row_count + 2


### iterating over the month of the specified year
for month in range(1,13):
    # get count of days for the actual month in for loop
    _, last_day_of_month = monthrange(actual_year, month)

    actual_cell = ws.cell(row=row_count,column=1,value=calendar.month_name[month])
    set_header_colomn_style(actual_cell)

    #for column in range(1,last_day_of_month + 1):
    column_position = 0
    for column in range(1,32):
        if column_position < last_day_of_month:
            actual_cell = ws.cell(row=row_count, column=column+1, value=column)
            set_header_colomn_style(actual_cell)
        else:
            actual_cell = ws.cell(row=row_count, column=column+1, value='')
            set_header_colomn_style(actual_cell)
        column_position += 1

    actual_cell = ws.cell(row=row_count,column=33,value=calendar.month_name[month])
    set_header_colomn_style(actual_cell)
    row_count = row_count + 1

    for user in users:
        ws.cell(row=row_count,column=1,value=user)
        ws.cell(row=row_count,column=33,value=user)
        row_count = row_count + 1

    # let two rows empty for spacing
    row_count = row_count + 2

### iterating over the first month of the year after
_, last_day_of_month = monthrange(actual_year + 1, 1)
actual_cell = ws.cell(row=row_count,column=1,value=calendar.month_name[1])
set_header_colomn_style(actual_cell)

column_position = 0
for column in range(1,32):
    if column_position < last_day_of_month:
        actual_cell = ws.cell(row=row_count, column=column+1, value=column)
        set_header_colomn_style(actual_cell)
    else:
        actual_cell = ws.cell(row=row_count, column=column+1, value='')
        set_header_colomn_style(actual_cell)
    column_position += 1
actual_cell = ws.cell(row=row_count,column=33,value=calendar.month_name[1])
set_header_colomn_style(actual_cell)
row_count = row_count + 1
for user in users:
    ws.cell(row=row_count,column=1,value=user)
    ws.cell(row=row_count,column=33,value=user)
    row_count = row_count + 1


# Save the file
wb.save(excel_file_name)
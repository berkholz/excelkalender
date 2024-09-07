"""
Script for creating an excel file as calendar overview. It contains all month of the specified year and each month before and after the year.

The script can be configured by the configuration file config.yml.
If you want a different file name, specify it in this file (variable config_name).

Holidays can be fetched from openapi.org via web request, see options oh_api_* in configuration file.
"""

############################ IMPORTS #####################
import sys

# import for creating excel files
from openpyxl import Workbook
import datetime
import calendar
from calendar import monthrange
from openpyxl.styles import PatternFill,Font, Color, Border, Side, Alignment

# import for pickung randomized elements from list colors[]
import random

# import for triggering REST API for holidays retrieval
import requests

# import for reading yaml configuration files
import yaml

# import for printing date in localized format
import locale

############################ VARIABLES #####################
excel_file_name = None

# row to start from, row_count must be greater equal 1
row_count = 1

# available colors for users, when not set via config file
colors = list()

# variable for configuration file settings/options
cal_config = None

# file name of the expected configuration file
config_name = 'config.yml'

cal = calendar.Calendar()

actual_year = None

# list with all (public and school) holidays as list for checking
holidays = list()

wb = Workbook()

############################ FUNCTIONS #####################
def load_config():
    """Load option from configfile and store the information in the variable cal_config with file name config_name."""
    global cal_config
    with open(config_name, 'r') as file:
        cal_config = yaml.safe_load(file)

def check_config_locale():
    """Function to check if the locale in the config file is set, otherise default de_DE is used."""
    global cal_config
    if not "locale" in cal_config:
        cal_config.update({ "locale" : "de_DE" })
        print("No locale found in confiuguration file {}, using default locale {}".format(config_name,cal_config['locale']))

def check_config_available_user_colors():
    """Function to check if available colors are set in the configuration file, otheriwse use default set of colors."""
    global cal_config
    global colors
    # check if colors are specified in configurations file, otherwise use default
    if "available_user_colors" in cal_config:
        colors = cal_config['available_user_colors'].copy()
    else:
        colors = ['00FF0000','0000FF00','000000FF','00FFFF00','00FF00FF','0000FFFF','00008000','00008080','00800080','009999FF','00FFFFCC','00FF8080','00CCCCFF','0099CC00','00FFCC00','00FF9900','00FF6600','00993300']
        print("Using default colors: {}".format(colors))

def check_config_excel_file_name():
    """Function to check if the output file name of the excel file is specified in the configuration file, otherwise use default kalender.xlsx."""
    global cal_config
    global excel_file_name
    if "excel_file_name" in cal_config:
        excel_file_name = cal_config['excel_file_name']
    else:
        excel_file_name = "kalender.xlsx"
        print("Using default file name for Excel output: {}".format(excel_file_name))

def check_config_year():
    """Function to check if a year different than the actual is specified. Default is actual year (datetime.date.today().year)."""
    global cal_config
    global actual_year
    if "year_for_calendar" in cal_config:
        actual_year = cal_config['year_for_calendar']
    else:
        actual_year = datetime.date.today().year
        print("No option \"year_for_calendar\" found in configuration file {}, using default: {}".format("config.yml", actual_year))

def check_config_header_bgcolor():
    global cal_config
    if "header_bgcolor" not in cal_config:
        cal_config.update( { "header_bgcolor" : "CFCFCF" } )
        print("No option \"header_bgcolor\" found in configuration file {}, using default: {}".format("config.yml", "CFCFCF"))

def check_config():
    """Global check function within subcheck functions for options from configuration file. The order of checking is relevant."""
    check_config_locale()
    check_config_available_user_colors()
    check_config_excel_file_name()
    check_config_year()
    check_config_header_bgcolor()

def print_list(list):
    """Function for printing lists."""
    print("[", end='')
    for element in list:
        print("{},".format(element), end='')
    print("]")

def associate_colors_to_users_if_not_set():
    """Check if users have color set in configuration file, otherwise use available_colors in configfile"""
    for user in cal_config['users']:
        if not "color" in user:
            key = random.sample(colors, 1)
            user.update( { "color" : key[0] } )
            colors.remove(key[0])

def add_title(worksheet):
    """Function to add a title row with year to the excel file."""
    global row_count
    worksheet.merge_cells('A1:AG1')
    worksheet.row_dimensions[1].height = 30
    cell = ws.cell(row=1,column=1,value=actual_year)
    cell.font = Font(name='Arial',size=16,bold=True)
    cell.alignment = Alignment(horizontal='center',vertical='center')
    cell.border = Border(bottom=Side(border_style="thin", color="000000"))
    #cell.fill = PatternFill(start_color="CFCFCF", end_color="0F0F0F", fill_type="solid")
    row_count = 3

def set_month_header_colomn_style (cell, bgcolor):
    """Function to set the style of the month header row, e.g. font face, font color, font name, border and cell fill color."""
    cell.font = Font(name='Arial', bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.border = Border(bottom=Side(border_style="thin", color="000000"), left=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"),top=Side(border_style="thin", color="000000"))
    if bgcolor == "":
        # if bgcolor empty the fill color is removed
        cell.fill = PatternFill()
    else:
        cell.fill = PatternFill(start_color=bgcolor, end_color=bgcolor, fill_type="solid")

def set_user_cell_style(cell, usercolor):
    """Function to set the style of the user rows, e.g. font face, font color, font name and cell fill color."""
    cell.font = Font(name='Arial', bold=False)
    cell.alignment = Alignment(horizontal='left')
    if usercolor == "":
        cell.fill = PatternFill()
    else:
        cell.fill = PatternFill(start_color=usercolor, end_color=usercolor, fill_type="solid")

def set_column_width(worksheet):
    """Function to set the width of the cells for all month and day coloumns."""
    # set cell width of month name
    column_range = ['A','AG']
    for column in column_range:
        worksheet.column_dimensions[column].width = 25

    # set cell width of days of month
    column_range = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF']
    for column in column_range:
        worksheet.column_dimensions[column].width = 5

def check_weekend(year, month, day):
    """Function to check if day, given as tripel(year, month, day) is a weekday."""
    day_to_check = datetime.date(year,month,day).isoweekday()

    # check if we have sunday (7) or saturday (6) and return accordingly color
    if day_to_check == 6:
        return True, 6
    elif day_to_check == 7:
        return True, 7
    else:
        return False, day_to_check

## TODO: remove function cause it works like get_user_color_for_weekday
def get_header_color_for_weekday(year,month,day):
    global cal_config
    bgcolor = cal_config['header_bgcolor']

    weekend, weekday = check_weekend(year,month,day)
    if weekend and weekday == 6:
        if "color_saturday" in cal_config:
           bgcolor = cal_config['color_saturday']
    elif weekend and weekday == 7:
        if "color_sunday" in cal_config:
           bgcolor = cal_config['color_sunday']
    return bgcolor

def get_user_color_for_weekday(year,month,day, default_user_color):
    global cal_config
    bgcolor = default_user_color

    weekend, weekday = check_weekend(year,month,day)
    if weekend and weekday == 6:
        if "color_saturday" in cal_config:
           bgcolor = cal_config['color_saturday']
    elif weekend and weekday == 7:
        if "color_sunday" in cal_config:
           bgcolor = cal_config['color_sunday']
    return bgcolor

def append_month(year, month_range_list):
    """This function is the main function for adding every month with users to the excel file."""
    global row_count
    global cal_config

    for month in month_range_list:
        ### MONTH TITLE ROW
        # add month name to begin of row
        _, last_day_of_month = monthrange(year, month)
        actual_cell = ws.cell(row=row_count,column=1,value=calendar.month_name[month])
        set_month_header_colomn_style(actual_cell, cal_config['header_bgcolor'])

        # Add days to same row of month name. That is the month line with its days.
        column_position = 0
        # We show 31 days, ...
        for column in range(1,32):
            # ..., but only days existing in month are filled with values.
            if column_position < last_day_of_month:
                actual_cell = ws.cell(row=row_count, column=column+1, value=column)
                set_month_header_colomn_style(actual_cell, get_user_color_for_weekday(year, month, column_position+1, cal_config['header_bgcolor']))
            else:
                actual_cell = ws.cell(row=row_count, column=column+1, value='')
                set_month_header_colomn_style(actual_cell, cal_config['header_bgcolor'])
            column_position += 1
        # add month name to end of row
        actual_cell = ws.cell(row=row_count,column=33,value=calendar.month_name[month])
        set_month_header_colomn_style(actual_cell, cal_config['header_bgcolor'])
        row_count = row_count + 1

        ### USER ROWS
        # add users as separate rows
        for user in cal_config['users']:
            # check if we have our holidays user
            if user['name'] == cal_config['oh_api_show_name']: # holiday user
                actual_cell = ws.cell(row=row_count,column=1,value=user['name'])
                set_user_cell_style(actual_cell, user['color'])

                for col in range(1,last_day_of_month + 1):
                    actual_cell = ws.cell(row=row_count,column=col+1,value='')
                    if is_date_in_holidays(datetime.datetime(year,month,col)):
                        set_user_cell_style(actual_cell, get_user_color_for_weekday(year,month,col,user['color']))
                    else:
                        set_user_cell_style(actual_cell, get_user_color_for_weekday(year,month,col,""))

                actual_cell = ws.cell(row=row_count,column=33,value=user['name'])
                set_user_cell_style(actual_cell, user['color'])
            else: # normal user
                actual_cell = ws.cell(row=row_count,column=1,value=user['name'])
                set_user_cell_style(actual_cell, user['color'])

                for col in range(1,last_day_of_month + 1):
                    actual_cell = ws.cell(row=row_count,column=col+1,value='')

                    if is_date_in_list_tupels(datetime.datetime(year,month,col), convert_str_date_list_to_datetime_list(user['vacation_dates'])):
                        set_user_cell_style(actual_cell, get_user_color_for_weekday(year,month,col,user['color']))
                    else:
                        set_user_cell_style(actual_cell, get_user_color_for_weekday(year,month,col,''))
                actual_cell = ws.cell(row=row_count,column=33,value=user['name'])
                set_user_cell_style(actual_cell, user['color'])
            row_count = row_count + 1
        # let two rows empty for spacing
        row_count = row_count + 2

def convert_str_date_list_to_datetime_list(str_date_list):
    """Function to convert list of string dates to datetime objects. This is used for comparing."""
    datetime_list = list()
    # iterate over string list with dates and store the datetime objects in the datetime_list
    for str_date_tupel in str_date_list:
        datetime_list.append(convert_str_date_to_datetime(str_date_tupel['startDate'],str_date_tupel['endDate']))

    return datetime_list

def convert_str_date_to_datetime(startDate, endDate):
    """Function to convert a date as string in format YYYY-MM-DD to a datetime object."""
    tmp_start = startDate.split("-")
    tmp_end = endDate.split('-')
    conv_startDate = datetime.datetime(int(tmp_start[0]),int(tmp_start[1]),int(tmp_start[2]))
    conv_endDate = datetime.datetime(int(tmp_end[0]),int(tmp_end[1]),int(tmp_end[2]))
    return conv_startDate,conv_endDate

def get_holidays(type_of_holidays):
    """Global functions to get the holidays from openapi.org via web request. The holidays to fetch are specified by parameter:
        public      = public holidays
        any other   = school holidays

    Result from opanapi.org is a json with dates as string representation.
    This will be converted to datetime obejcts and add to the global variable holidays.
    """
    global holidays
    global cal_config
    if type_of_holidays == "public":
        query_string = "PublicHolidays"
    else:
        query_string = "SchoolHolidays"
    url = cal_config['oh_api_base_url'] + query_string + "?countryIsoCode=" + cal_config['oh_api_country_iso_code'] + "&languageIsoCode=" + cal_config['oh_api_language_iso_code'] + "&validFrom=" + str(actual_year - 1) + "-12-01" + "&validTo=" + str(actual_year + 1) + "-01-31" + "&subdivisionCode=" + cal_config['oh_api_subdivision_code']

    response = requests.get(url)
    response_json = response.json()

    for holiday in response_json:
        holidays.append(convert_str_date_to_datetime(holiday['startDate'],holiday['endDate']))

def get_public_holidays():
    """Meta function to get public holidays by calling the function get_holidays("public")."""
    get_holidays("public")

def get_school_holidays():
    """Meta function to get school holidays by calling the function get_holidays("school")."""
    get_holidays("school")

def add_holidays_as_user():
    """Function to add holidays as user on last position. This user is added to excel file."""
    new_user = dict({ "name" : cal_config['oh_api_show_name'], "color" : cal_config['oh_api_show_color']})
    cal_config['users'].append(new_user)

def print_holidays():
    for tupel in holidays:
        print(tupel)

def is_date_in_holidays(date):
    """Function to check if the given date by parameter is in the holidays."""
    global holidays
    return is_date_in_list_tupels(date, holidays)

def is_date_in_list_tupels(date, list_of_date_tupels):
    for start, end in list_of_date_tupels:
        if date >= start and date <= end:
            return True
    return False

############################ MAIN #####################

load_config()

check_config()

locale.setlocale(locale.LC_ALL, cal_config['locale'])

# show holidays only if specified, default is True
if cal_config['oh_api_show_holidays']:
    get_school_holidays()
    add_holidays_as_user()

# associate colors to users
associate_colors_to_users_if_not_set()

# grab the active worksheet
ws = wb.active
# set title of the worksheet
ws.title = "Urlaubsplanung " + str(actual_year)

# add header (year) as title to the excel
add_title(ws)

# set colomn width of days
set_column_width(ws)

# generating last month of year before
append_month(actual_year - 1,list([12]))
# generating specified year
append_month(actual_year,range(1,13))
# generating first month of year after
append_month(actual_year + 1,list([1]))

# Save the file
wb.save(excel_file_name)
